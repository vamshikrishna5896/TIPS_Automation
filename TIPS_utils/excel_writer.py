from openpyxl import load_workbook
from pathlib import Path


class ExcelWriter:

    # =====================================
    # Dental Policy Update
    # Match using MyKadID
    # Policy Number -> Column 3
    # =====================================
    @staticmethod
    def update_dental_policy_number(sheet_name, mykad,quote_number, policy_number):

        project_root = Path(__file__).parent.parent
        excel_file = project_root / "testdata_excel" / "test_data.xlsx"

        workbook = load_workbook(excel_file)
        sheet = workbook[sheet_name]

        for row in range(2, sheet.max_row + 1):

            excel_mykad = str(sheet.cell(row=row, column=1).value)

            if excel_mykad == str(mykad):

                sheet.cell(row=row, column=3).value = quote_number

                sheet.cell(row=row, column=4).value = policy_number
                break

        workbook.save(excel_file)

        print(f"Dental Quote Number {quote_number} updated successfully in {sheet_name}")

        print(f"Dental Policy Number {policy_number} updated successfully in {sheet_name}")

    # =====================================
    # PA Endorsement / Correction
    # Policy Number -> Column 1
    # First Empty Row
    # =====================================
    @staticmethod
    def update_pa_policy_number(sheet_name, policy_number):

        project_root = Path(__file__).parent.parent
        excel_file = project_root / "testdata_excel" / "test_data.xlsx"

        workbook = load_workbook(excel_file)
        sheet = workbook[sheet_name]

        for row in range(2, sheet.max_row + 2):

            if sheet.cell(row=row, column=1).value in [None, ""]:

                sheet.cell(row=row, column=1).value = policy_number
                break

        workbook.save(excel_file)

        print(f"PA Policy Number {policy_number} updated successfully in {sheet_name}")

    # =====================================
    # PC_RV Motor Quote & policy Number Update
    # =====================================
    @staticmethod
    def update_motor_pc_rv_quote_number(sheet_name, reg_no, quote_number, policy_number):

        project_root = Path(__file__).parent.parent

        excel_file = project_root / "testdata_excel" / "test_data.xlsx"

        workbook = load_workbook(excel_file)

        sheet = workbook[sheet_name]

        for row in range(2, sheet.max_row + 1):

            excel_reg_no = str(sheet.cell(row=row, column=1).value)

            if excel_reg_no == str(reg_no):
                # Column 5 = Quote Number
                sheet.cell(row=row, column=5).value = quote_number
                # Column 6 = Policy Number
                sheet.cell(row=row, column=6).value = policy_number

                break

        workbook.save(excel_file)

        print(f"Quote Number {quote_number} updated successfully in {sheet_name}")
        print(f"Motor Policy Number {policy_number} updated successfully in {sheet_name}")

    # =====================================
    # MC_RV Motor Quote & policy Number Update
    # =====================================
    @staticmethod
    def update_motor_mc_rv_policy_number(sheet_name, reg_no,quote_number, policy_number):

        project_root = Path(__file__).parent.parent
        excel_file = project_root / "testdata_excel" / "test_data.xlsx"

        workbook = load_workbook(excel_file)
        sheet = workbook[sheet_name]

        for row in range(2, sheet.max_row + 1):

            excel_reg_no = str(sheet.cell(row=row, column=1).value)

            if excel_reg_no == str(reg_no):
                # Column 5 = Quote Number
                sheet.cell(row=row, column=5).value = quote_number
                # Column 6 = Policy Number
                sheet.cell(row=row, column=6).value = policy_number
                break

        workbook.save(excel_file)
        print(f"Quote Number {quote_number} updated successfully in {sheet_name}")
        print(f"Motor Policy Number {policy_number} updated successfully in {sheet_name}")

    # =====================================
    # PC_NV Motor Quote & policy Number Update
    # =====================================
    @staticmethod
    def update_motor_pc_nv_details(sheet_name, mykad, quote_number, policy_number):

        project_root = Path(__file__).parent.parent

        excel_file = (project_root / "testdata_excel" / "test_data.xlsx")

        workbook = load_workbook(excel_file)

        sheet = workbook[sheet_name]

        for row in range(2, sheet.max_row + 1):

            excel_mykad = str(sheet.cell(row=row, column=1).value)

            if excel_mykad == str(mykad):
                # Quote Number
                sheet.cell(row=row, column=5).value = quote_number
                # Policy Number
                sheet.cell(row=row, column=6).value = policy_number

                break

        workbook.save(excel_file)

        print(f"Quote Number : {quote_number}")

        print(f"Policy Number : {policy_number}")

