from openpyxl import load_workbook
from pathlib import Path


class ExcelReader:

    @staticmethod
    def get_test_data(sheet_name):

        project_root = Path(__file__).parent.parent

        excel_file = (
            project_root
            / "testdata_excel"
            / "test_data.xlsx"
        )

        workbook = load_workbook(excel_file)

        print("Available Sheets:", workbook.sheetnames)

        sheet = workbook[sheet_name]

        headers = []

        for cell in sheet[1]:
            headers.append(cell.value)

        data = []

        for row in sheet.iter_rows(
                min_row=2,
                values_only=True):

            row_data = {}

            for header, value in zip(headers, row):
                row_data[header] = value

            data.append(row_data)

        return data